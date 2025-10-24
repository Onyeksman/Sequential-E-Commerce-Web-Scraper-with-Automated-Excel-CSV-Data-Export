import scrapy
from scrapy.crawler import CrawlerProcess
import csv
import os
import re
from ftfy import fix_text
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from pathlib import Path

class ShopSpider(scrapy.Spider):
    name = "shop_spider"
    start_urls = ["https://scrapeme.live/shop/"]

    custom_settings = {
        "LOG_LEVEL": "INFO",
        "AUTOTHROTTLE_ENABLED": False,
        "CONCURRENT_REQUESTS": 1,   # ensure strict sequential crawling
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        today = datetime.now().strftime("%Y-%m-%d")
        self.items = OrderedDict()
        self.seq = 0
        self.currency_symbol = "$"
        self.csv_file = Path(f"pokemon_{today}.csv")
        self.xlsx_file = Path(f"pokemon_{today}.xlsx")

    # ---------- helpers ----------
    def clean_text(self, text):
        """Normalize whitespace and fix encoding glitches."""
        if not text:
            return ""
        return fix_text(re.sub(r"\s+", " ", text).strip())

    def parse_price(self, price_block):
        """Extract numeric price and display string with $ currency."""
        if not price_block:
            return "N/A", None

        currency = price_block.css("span.woocommerce-Price-currencySymbol::text").get(default="").strip()
        amount_text_nodes = price_block.xpath("text()").getall()
        amount_text = " ".join([a for a in amount_text_nodes]).strip() or price_block.get() or ""
        currency = fix_text(currency or self.currency_symbol)
        amount_text = fix_text(amount_text)

        m = re.search(r"[\d\.,]+", amount_text)
        if not m:
            return "N/A", None

        num_str = m.group(0).replace(",", "")
        try:
            num = float(num_str)
        except Exception:
            return "N/A", None

        display = f"{self.currency_symbol}{num:,.2f}"
        return display, num

    # ---------- parse ----------
    def parse(self, response):
        self.logger.info(f"Scraping page: {response.url}")
        products = response.css("li.product.type-product")
        self.logger.info(f"Found {len(products)} products on {response.url}")

        for product in products:
            self.seq += 1
            seq = self.seq

            name = self.clean_text(product.css("h2.woocommerce-loop-product__title::text").get(default=""))
            price_block = product.css("span.woocommerce-Price-amount.amount")
            price_display, price_num = self.parse_price(price_block)
            image = self.clean_text(product.css("img.attachment-woocommerce_thumbnail::attr(src)").get(default=""))

            record = {
                "Name": name or "",
                "PriceDisplay": price_display if price_display else "",
                "PriceNumeric": price_num,
                "Image": image or "",
            }

            self.items[seq] = record

        # Pagination: sequential follow-up
        next_page = response.css("a.next.page-numbers::attr(href), a.next::attr(href)").get()
        if next_page:
            self.logger.info(f"➡️ Proceeding to next page: {next_page}")
            yield scrapy.Request(
                url=response.urljoin(next_page),
                callback=self.parse,
                priority=-1,
                dont_filter=True,
            )

    # ---------- finalize ----------
    def closed(self, reason):
        keys_csv = ["Name", "Price", "Image"]
        cleaned = []
        seen = set()

        for seq in self.items.keys():
            it = self.items[seq]
            name = self.clean_text(it.get("Name", "")) or "N/A"
            image = self.clean_text(it.get("Image", "")) or "N/A"
            price_display = it.get("PriceDisplay") or "N/A"
            price_num = it.get("PriceNumeric")

            row_tuple = (name, price_display, image)
            if row_tuple in seen:
                continue
            seen.add(row_tuple)

            cleaned.append({
                "Name": name,
                "Price": price_display,
                "PriceNumeric": price_num,
                "Image": image,
            })

        # remove entirely empty rows
        filtered = [
            r for r in cleaned
            if not all((v == "N/A" or v == "" or v is None) for v in (r["Name"], r["Price"], r["Image"]))
        ]

        # Preserve exact order (no sorting)
        filtered = list(filtered)

        # Ensure overwrite safety
        for file in [self.csv_file, self.xlsx_file]:
            if file.exists():
                os.remove(file)

        # --- Write CSV ---
        with open(self.csv_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=keys_csv, quoting=csv.QUOTE_ALL)
            writer.writeheader()
            for r in filtered:
                writer.writerow({
                    "Name": r["Name"] or "N/A",
                    "Price": r["Price"] or "N/A",
                    "Image": r["Image"] or "N/A",
                })

        # --- Excel formatting ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="medium"),
            right=Side(style="medium"),
            top=Side(style="medium"),
            bottom=Side(style="medium"),
        )
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        na_font = Font(color="808080", italic=True)

        headers = ["Name", "Price", "Image"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        for r_idx, r in enumerate(filtered, start=2):
            cell_name = ws.cell(row=r_idx, column=1, value=r["Name"] or "N/A")
            cell_name.border = border
            if r_idx % 2 == 0:
                cell_name.fill = alt_fill
            if cell_name.value == "N/A":
                cell_name.font = na_font

            if r["PriceNumeric"] is not None:
                cell_price = ws.cell(row=r_idx, column=2, value=r["PriceNumeric"])
                cell_price.number_format = f'{self.currency_symbol}#,##0.00'
            else:
                cell_price = ws.cell(row=r_idx, column=2, value="N/A")
                cell_price.font = na_font

            cell_price.border = border
            if r_idx % 2 == 0:
                cell_price.fill = alt_fill

            cell_img = ws.cell(row=r_idx, column=3, value=r["Image"] or "N/A")
            cell_img.border = border
            if r_idx % 2 == 0:
                cell_img.fill = alt_fill
            if cell_img.value == "N/A":
                cell_img.font = na_font

        # --- Auto-fit columns ---
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                text = f"{self.currency_symbol}{cell.value:,.2f}" if cell.column == 2 and isinstance(cell.value, (int, float)) else str(cell.value or "")
                max_length = max(max_length, len(text) * 1.2)
            ws.column_dimensions[col_letter].width = max(12, min(max_length, 80))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        meta_row = len(filtered) + 3
        note = f"📊 Sourced from (https://scrapeme.live/shop/) — {datetime.now():%Y-%m-%d %H:%M:%S}"
        meta_cell = ws.cell(row=meta_row, column=1, value=note)
        meta_cell.font = Font(color="808080", italic=True)
        ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=len(headers))

        wb.save(self.xlsx_file)

        self.logger.info(f"Wrote {len(filtered)} products sequentially to {self.csv_file.resolve()} and {self.xlsx_file.resolve()} ({reason})")

if __name__ == "__main__":
    process = CrawlerProcess()
    process.crawl(ShopSpider)
    process.start()
